#!/usr/bin/env python3
"""cafe_reviews.csv -> VoC 분석 엑셀 (별점 분포 차트 + 테마 피벗 + 부정 Top3).
실행: python3 make_voc_xlsx.py
산출: out/voc_analysis.xlsx
"""
import csv, os
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "cafe_reviews.csv")
OUT_DIR = os.path.join(HERE, "out")
os.makedirs(OUT_DIR, exist_ok=True)
OUT = os.path.join(OUT_DIR, "voc_analysis.xlsx")

rows = []
with open(SRC, encoding="utf-8") as f:
    for r in csv.DictReader(f):
        r["rating"] = int(r["rating"])
        rows.append(r)

# ---- 집계 ----
total = len(rows)
avg = round(sum(r["rating"] for r in rows) / total, 2)
rating_dist = Counter(r["rating"] for r in rows)              # 별점 분포
theme_cnt = Counter(r["theme"] for r in rows)                 # 전체 테마
NEG = [r for r in rows if r["rating"] <= 2]                   # 부정(별점<=2)
neg_theme = Counter(r["theme"] for r in NEG)                  # 부정 테마

# ---- 스타일 ----
HEAD = Font(bold=True, color="FFFFFF", size=11)
HEAD_FILL = PatternFill("solid", fgColor="2F5496")
TITLE = Font(bold=True, size=14, color="2F5496")
NEG_FILL = PatternFill("solid", fgColor="FCE4E4")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center")

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD; cell.fill = HEAD_FILL
        cell.alignment = CENTER; cell.border = BORDER

wb = Workbook()

# ===== Sheet 1: 요약 대시보드 =====
ws = wb.active
ws.title = "요약"
ws["A1"] = "☕ 데일리브루 고객 리뷰(VoC) 분석"
ws["A1"].font = TITLE
ws["A2"] = f"리뷰 {total}건 · 평균 별점 {avg} · 부정(★≤2) {len(NEG)}건"
ws["A2"].font = Font(italic=True, color="595959")

ws["A4"] = "핵심 인사이트"
ws["A4"].font = Font(bold=True, size=12)
top_neg = neg_theme.most_common(1)[0]
insights = [
    f"1) 부정 리뷰 1위 테마 = '{top_neg[0]}' ({top_neg[1]}건 / 부정 {len(NEG)}건)",
    f"2) 평균 별점 {avg} — 맛/친절은 호평, 운영(대기·좌석)에서 이탈 발생",
    "3) '대기시간' 불만은 주말 피크에 집중 → 좌석 회전·예약/주문 동선 개선 필요",
]
for i, t in enumerate(insights):
    ws.cell(row=5 + i, column=1, value=t)

ws.column_dimensions["A"].width = 60

# ===== Sheet 2: 원천 데이터 =====
ws2 = wb.create_sheet("리뷰원본")
cols = ["date", "platform", "rating", "review", "theme"]
ws2.append(["날짜", "플랫폼", "별점", "리뷰", "테마"])
style_header(ws2, 1, len(cols))
for r in rows:
    ws2.append([r["date"], r["platform"], r["rating"], r["review"], r["theme"]])
    if r["rating"] <= 2:
        for c in range(1, len(cols) + 1):
            ws2.cell(row=ws2.max_row, column=c).fill = NEG_FILL
widths = [12, 9, 7, 50, 11]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:E{ws2.max_row}"

# ===== Sheet 3: 별점 분포 (+ 막대/원형 차트) =====
ws3 = wb.create_sheet("별점분포")
ws3.append(["별점", "건수", "비율(%)"])
style_header(ws3, 1, 3)
for star in [5, 4, 3, 2, 1]:
    cnt = rating_dist.get(star, 0)
    ws3.append([f"★{star}", cnt, round(cnt / total * 100, 1)])
for col in "ABC":
    ws3.column_dimensions[col].width = 12

bar = BarChart(); bar.title = "별점 분포"; bar.type = "col"; bar.style = 10
bar.y_axis.title = "건수"; bar.x_axis.title = "별점"
data = Reference(ws3, min_col=2, min_row=1, max_row=6)
cats = Reference(ws3, min_col=1, min_row=2, max_row=6)
bar.add_data(data, titles_from_data=True); bar.set_categories(cats)
bar.legend = None; bar.height = 7; bar.width = 12
ws3.add_chart(bar, "E2")

pie = PieChart(); pie.title = "별점 비율"
pdata = Reference(ws3, min_col=2, min_row=1, max_row=6)
pie.add_data(pdata, titles_from_data=True); pie.set_categories(cats)
pie.height = 7; pie.width = 9
ws3.add_chart(pie, "E18")

# ===== Sheet 4: 테마 피벗 (전체 vs 부정) =====
ws4 = wb.create_sheet("테마피벗")
ws4.append(["테마", "전체 건수", "부정 건수", "부정 비중(%)"])
style_header(ws4, 1, 4)
for theme, cnt in theme_cnt.most_common():
    ncnt = neg_theme.get(theme, 0)
    ws4.append([theme, cnt, ncnt, round(ncnt / cnt * 100, 0) if cnt else 0])
for col in "ABCD":
    ws4.column_dimensions[col].width = 13

tbar = BarChart(); tbar.title = "테마별 리뷰 (전체/부정)"; tbar.type = "col"; tbar.style = 12
tdata = Reference(ws4, min_col=2, max_col=3, min_row=1, max_row=ws4.max_row)
tcats = Reference(ws4, min_col=1, min_row=2, max_row=ws4.max_row)
tbar.add_data(tdata, titles_from_data=True); tbar.set_categories(tcats)
tbar.height = 8; tbar.width = 14
ws4.add_chart(tbar, "F2")

# ===== Sheet 5: 부정 리뷰 Top3 테마 + 액션 =====
ws5 = wb.create_sheet("부정Top3")
ws5["A1"] = "🔴 부정 리뷰(★≤2) Top3 테마 & 개선 액션"
ws5["A1"].font = TITLE
ws5.append([])
ws5.append(["순위", "테마", "건수", "대표 리뷰", "개선 액션(제안)"])
style_header(ws5, 3, 5)
ACTION = {
    "대기시간": "주말 피크 시간대 예약/모바일 주문 도입, 좌석 회전 동선 개선",
    "가격": "가성비 세트(디저트+음료) 구성, 가격 가치 커뮤니케이션 강화",
    "청결": "피크 후 테이블 정리 주기 단축, 셀프 반납대 운영",
    "맛": "신메뉴 테스트 및 레시피 표준화",
    "분위기": "좌석 간격·소음 관리, 카공존 분리",
    "친절": "응대 매뉴얼/온보딩 보강",
}
for rank, (theme, cnt) in enumerate(neg_theme.most_common(3), 1):
    rep = next((r["review"] for r in NEG if r["theme"] == theme), "")
    ws5.append([rank, theme, cnt, rep, ACTION.get(theme, "")])
ws5.column_dimensions["A"].width = 6
ws5.column_dimensions["B"].width = 11
ws5.column_dimensions["C"].width = 7
ws5.column_dimensions["D"].width = 45
ws5.column_dimensions["E"].width = 50
for row in ws5.iter_rows(min_row=4, max_row=ws5.max_row):
    for c in row:
        c.alignment = Alignment(vertical="center", wrap_text=True)

wb.save(OUT)
print("saved:", OUT)
print("avg:", avg, "| neg total:", len(NEG))
print("neg theme rank:", neg_theme.most_common())
